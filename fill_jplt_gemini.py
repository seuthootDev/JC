#!/usr/bin/env python3
"""
JPLT.xlsx -> Google Gemini API -> jplt_vocabulary.json (시트별 DAY 01~50)

사용법:
  pip install google-generativeai openpyxl
  set GEMINI_API_KEY=your-key
  python fill_jplt_gemini.py

  python fill_jplt_gemini.py --sheet "DAY 03"
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

try:
    import google.generativeai as genai
except ImportError:
    print(
        "google-generativeai 패키지가 없습니다: pip install google-generativeai",
        file=sys.stderr,
    )
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "JPLT.xlsx"
DEFAULT_JSON = SCRIPT_DIR / "jplt_vocabulary.json"
DEFAULT_SKIP = frozenset({"DAY 00"})
DEFAULT_MODEL = "gemini-2.5-flash"

SYSTEM_PROMPT = """\
당신은 일본어-한국어 학습용 사전 편집자입니다.
사용자가 주는 일본어 단어 목록에 대해 한국어 뜻만 채웁니다.

규칙:
1. "뜻"은 한글(가-힣)만 — 최소 1글자 이상 반드시 포함.
2. 뜻에 일본어 한자·가나·로마자·영어 금지. (すし→초밥 O, 寿司 X / もも→복숭아 O, 桃 X)
3. 요미가나·단어를 그대로 복사하지 마세요. 한국어 의미만.
4. 숫자 한자(一, 二…)는 "일, 이, 삼".
5. 맥락상 가장 흔한 JLPT/일상 의미 하나만.
6. 출력: {"entries": [ ... ]} JSON만.
7. entries 개수·순서는 입력과 동일. 번호·단어·요미가나는 입력값 그대로.
"""

RETRY_HINT = (
    "\n\n[재요청] 이전 답에 한자/일본어가 들어갔습니다. "
    '"뜻"은 한글만. 예: すし→초밥, もも→복숭아, 天ぷら→튀김, 桃·天婦羅 금지.'
)


@dataclass(frozen=True)
class SheetColumns:
    word: int
    meaning: int
    num: int | None = None
    reading: int | None = None


def load_env_file(path: Path) -> None:
    """.env.local 등에서 KEY=VALUE 로드 (이미 설정된 변수는 덮어쓰지 않음)."""
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key, value = key.strip(), value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Gemini로 JPLT 단어장 JSON 생성")
    p.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    p.add_argument("--json-output", type=Path, default=DEFAULT_JSON)
    p.add_argument("--api-key", default=None, help="미지정 시 GEMINI_API_KEY")
    p.add_argument(
        "--model",
        default=DEFAULT_MODEL,
        help="기본 gemini-2.5-flash (구 2.0-flash는 신규 계정 불가)",
    )
    p.add_argument("--sheet", default=None, help='예: "DAY 01"')
    p.add_argument("--delay", type=float, default=1.0, help="시트 간 대기(초)")
    p.add_argument("--update-xlsx", action="store_true", help="엑셀 뜻 열도 갱신")
    p.add_argument("--force", action="store_true", help="이미 채운 시트도 다시 요청")
    p.add_argument("--dry-run", action="store_true")
    return p.parse_args()


def sheet_sort_key(name: str) -> tuple[int, str]:
    m = re.search(r"(\d+)", name)
    return (int(m.group(1)) if m else 9999, name)


def detect_columns(ws) -> SheetColumns | None:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        return None
    labels = [str(c).strip() if c is not None else "" for c in header]

    def find(*names: str) -> int | None:
        for i, label in enumerate(labels):
            if label in names:
                return i
        return None

    word_i = find("단어")
    meaning_i = find("뜻")
    if word_i is None:
        return None
    return SheetColumns(
        word=word_i,
        meaning=meaning_i if meaning_i is not None else -1,
        num=find("번호"),
        reading=find("요미가나"),
    )


def normalize_reading(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.startswith("="):
        return None
    return s


def has_hangul(text: str) -> bool:
    return any("\uac00" <= c <= "\ud7a3" for c in text)


def is_good_meaning(meaning: str, word: str) -> bool:
    if not meaning or not meaning.strip():
        return False
    m = meaning.strip()
    if not has_hangul(m):
        return False
    if m == word:
        return False
    jp = sum(1 for c in m if "\u3040" <= c <= "\u30ff" or "\u4e00" <= c <= "\u9fff")
    if jp > len(m) // 2:
        return False
    return True


def read_sheet_entries(ws, cols: SheetColumns) -> list[dict[str, Any]]:
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= cols.word:
            continue
        word_cell = row[cols.word]
        if word_cell is None or not str(word_cell).strip():
            continue
        word = str(word_cell).strip()
        if word.isdigit():
            continue
        num = row[cols.num] if cols.num is not None and len(row) > cols.num else None
        reading = (
            normalize_reading(row[cols.reading])
            if cols.reading is not None and len(row) > cols.reading
            else None
        )
        entries.append(
            {"번호": num, "단어": word, "요미가나": reading, "뜻": None}
        )
    return entries


def sheet_needs_work(
    sheet_name: str, entries: list[dict], vocabulary: dict, force: bool
) -> bool:
    if force:
        return True
    existing = vocabulary.get(sheet_name, [])
    if len(existing) != len(entries):
        return True
    for e in existing:
        if not is_good_meaning(str(e.get("뜻") or ""), e.get("단어", "")):
            return True
    return False


def parse_llm_json_text(text: str) -> Any:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    return json.loads(text)


def parse_llm_json_entries(text: str) -> list[dict]:
    data = parse_llm_json_text(text)
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and "entries" in data:
        return data["entries"]
    raise ValueError("JSON에 entries 배열이 없습니다")


def merge_validated_entries(
    sheet_name: str,
    sources: list[dict[str, Any]],
    parsed: list[dict],
) -> list[dict[str, Any]]:
    if len(parsed) != len(sources):
        raise ValueError(
            f"{sheet_name}: 항목 수 불일치 (기대 {len(sources)}, 받음 {len(parsed)})"
        )
    out = []
    for src, item in zip(sources, parsed):
        word = src["단어"]
        meaning = str(item.get("뜻") or "").strip()
        if not is_good_meaning(meaning, word):
            raise ValueError(f"{sheet_name} {word}: 뜻 이상 -> {meaning!r}")
        out.append(
            {
                "번호": src["번호"],
                "단어": word,
                "요미가나": src["요미가나"],
                "뜻": meaning,
            }
        )
    return out


def gemini_fill_sheet(
    model: genai.GenerativeModel,
    sheet_name: str,
    entries: list[dict[str, Any]],
    max_attempts: int = 3,
) -> list[dict[str, Any]]:
    payload = [
        {"번호": e["번호"], "단어": e["단어"], "요미가나": e["요미가나"]}
        for e in entries
    ]
    base_msg = (
        f"시트: {sheet_name}\n"
        f"아래 {len(payload)}개 항목에 뜻을 채워 JSON으로 답하세요.\n\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )
    user_msg = base_msg
    last_err: ValueError | None = None

    for attempt in range(max_attempts):
        response = model.generate_content(user_msg)
        raw = response.text or ""
        parsed = parse_llm_json_entries(raw)
        try:
            return merge_validated_entries(sheet_name, entries, parsed)
        except ValueError as e:
            last_err = e
            if attempt + 1 >= max_attempts:
                raise
            user_msg = base_msg + RETRY_HINT + f"\n\n오류: {e}"
            time.sleep(1.0)

    raise last_err or ValueError(f"{sheet_name}: 번역 실패")


def is_resumable_api_error(exc: BaseException) -> bool:
    msg = str(exc).lower()
    return any(
        k in msg
        for k in (
            "rate",
            "limit",
            "quota",
            "429",
            "resource exhausted",
            "too many requests",
            "tokens",
        )
    )


def save_json(path: Path, data: Any) -> None:
    tmp = path.with_suffix(".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    tmp.replace(path)


def apply_to_xlsx(
    wb: openpyxl.Workbook, sheet_name: str, entries: list[dict], cols: SheetColumns
) -> None:
    if cols.meaning < 0:
        return
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    ei = 0
    for row_idx, row in enumerate(rows, start=2):
        if not row or len(row) <= cols.word:
            continue
        if row[cols.word] is None or not str(row[cols.word]).strip():
            continue
        if str(row[cols.word]).strip().isdigit():
            continue
        if ei >= len(entries):
            break
        ws.cell(row=row_idx, column=cols.meaning + 1, value=entries[ei]["뜻"])
        ei += 1


def load_vocabulary(path: Path) -> dict[str, list[dict]]:
    if not path.exists():
        return {}
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def build_gemini_model(api_key: str, model_name: str) -> genai.GenerativeModel:
    genai.configure(api_key=api_key)
    return genai.GenerativeModel(
        model_name=model_name,
        system_instruction=SYSTEM_PROMPT,
        generation_config=genai.GenerationConfig(
            temperature=0.1,
            response_mime_type="application/json",
        ),
    )


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass

    load_env_file(SCRIPT_DIR / ".env.local")
    load_env_file(SCRIPT_DIR / ".env")

    args = parse_args()
    api_key = "AIzaSyDnJsY_xJpHpEgYFJgSkWsKg7zV_JDrzps"
    if not args.dry_run and not api_key:
        print(
            "GEMINI_API_KEY 환경 변수 또는 --api-key 필요\n"
            "  https://aistudio.google.com/apikey",
            file=sys.stderr,
        )
        return 1

    input_path = args.input.resolve()
    json_path = args.json_output.resolve()
    if not input_path.exists():
        print(f"파일 없음: {input_path}", file=sys.stderr)
        return 1

    print(f"입력: {input_path}")
    print(f"JSON: {json_path}")
    print(f"모델: {args.model}")

    vocabulary = load_vocabulary(json_path)
    wb = openpyxl.load_workbook(input_path)
    model = None if args.dry_run else build_gemini_model(api_key, args.model)

    sheets = sorted(
        [n for n in wb.sheetnames if n not in DEFAULT_SKIP],
        key=sheet_sort_key,
    )
    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(f"시트 없음: {args.sheet}", file=sys.stderr)
            return 1
        sheets = [args.sheet]

    todo_sheets = []
    for name in sheets:
        cols = detect_columns(wb[name])
        if cols is None:
            print(f"  [{name}] 헤더 없음, 건너뜀")
            continue
        entries = read_sheet_entries(wb[name], cols)
        if not entries:
            continue
        if sheet_needs_work(name, entries, vocabulary, args.force):
            todo_sheets.append((name, cols, entries))
        else:
            print(f"  [{name}] 이미 완료 ({len(entries)}개)")

    total_words = sum(len(e) for _, _, e in todo_sheets)
    print(f"처리 대상: {len(todo_sheets)}개 시트, {total_words}개 단어")

    if args.dry_run:
        wb.close()
        return 0

    print("모드: DAY 01부터 시트별 순차 (완료 시트 건너뜀, 한도 걸리면 저장 후 종료)")

    done = 0
    stopped_early = False
    script = "fill_jplt_gemini.py"
    try:
        if not todo_sheets:
            print("처리할 시트 없음 (전부 완료됐거나 --force 로 재실행).")
        else:
            for sheet_name, cols, entries in todo_sheets:
                print(f"  [{sheet_name}] Gemini 요청 중 ({len(entries)}개)...")
                try:
                    filled = gemini_fill_sheet(model, sheet_name, entries)
                except ValueError as e:
                    save_json(json_path, vocabulary)
                    if args.update_xlsx:
                        wb.save(input_path)
                    print(
                        f"\n[검증 실패] {sheet_name}: {e}\n"
                        f"  이 시트는 건너뜁니다. 나중에:\n"
                        f'  python {script} --sheet "{sheet_name}"'
                    )
                    continue
                except Exception as e:
                    save_json(json_path, vocabulary)
                    if args.update_xlsx:
                        wb.save(input_path)
                    if is_resumable_api_error(e):
                        print(
                            f"\n[한도/할당량] {sheet_name} 처리 중 중단\n"
                            f"  {e}\n"
                            f"  JSON 저장됨. 잠시 후 같은 명령으로 이어집니다:\n"
                            f"  python {script}"
                        )
                        stopped_early = True
                        break
                    raise

                vocabulary[sheet_name] = filled
                save_json(json_path, vocabulary)
                if args.update_xlsx:
                    apply_to_xlsx(wb, sheet_name, filled, cols)
                    wb.save(input_path)
                done += 1
                sample = filled[0]
                print(
                    f"    완료 — 예: {sample['단어']} -> {sample['뜻']} "
                    f"({done}/{len(todo_sheets)}시트)"
                )
                if args.delay > 0 and done < len(todo_sheets):
                    time.sleep(args.delay)
    except KeyboardInterrupt:
        print(f"\n[중단] JSON 저장됨. python {script} 로 이어서 실행.")
        stopped_early = True
    except Exception as e:
        save_json(json_path, vocabulary)
        if args.update_xlsx:
            wb.save(input_path)
        wb.close()
        if is_resumable_api_error(e):
            print(f"\n[한도/할당량] {e}\nJSON 저장됨. python {script} 로 이어서 실행.")
            return 0
        print(f"\n[오류] {e}", file=sys.stderr)
        return 2
    finally:
        try:
            if args.update_xlsx:
                wb.save(input_path)
        except Exception:
            pass
        wb.close()

    total = sum(len(v) for v in vocabulary.values())
    filled_sheets = len(vocabulary)
    if stopped_early:
        print(f"\n중단됨 — JSON {total}항목, {filled_sheets}개 시트 -> {json_path}")
    else:
        print(f"\n완료 — JSON {total}항목, {filled_sheets}개 시트 -> {json_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
